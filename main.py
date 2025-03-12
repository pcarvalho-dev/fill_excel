import re
import openpyxl
from openpyxl.styles import Font


def extract_data(text):
    """Extracts specific fields from the given text."""
    data = {}

    data['Cliente'] = re.search(r'Cliente:\s*(.+)', text)
    data['Protocolo'] = re.search(r'Protocolo:\s*([\w-]+)', text)
    data['Valor Total'] = re.search(r'Valor Total:\s*R\$ ([\d,.]+)', text)
    data['KM'] = re.search(r'KM:\s*([\d,.]+)', text)

    # Extract matched values or set to empty string if not found
    for key in data:
        data[key] = data[key].group(1) if data[key] else ''

    return data


def find_insert_row(sheet):
    """Finds the first empty row before the 'Total' row without overwriting existing data."""
    total_row = None
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if any(cell and isinstance(cell, str) and "Total" in cell for cell in row):
            total_row = row_idx
            break

    if total_row:
        return total_row  # Insert as a new row above 'Total' row

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None for cell in row):
            return row_idx  # Insert at first empty row

    return sheet.max_row + 1  # Default to last row if no empty row is found


def write_to_excel(data, filename="output.xlsx"):
    """Writes extracted data to an existing Excel file, ensuring it aligns with headers and applies font size 8."""
    try:
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active
    except FileNotFoundError:
        print("File not found. Please check the file path and try again.")
        return

    # Get headers from the first row
    headers = [cell.value for cell in sheet[1]]

    # Find correct column indexes
    col_cliente = headers.index("Cliente:") + \
        1 if "Cliente:" in headers else None
    col_protocolo = headers.index(
        "Protocolo:") + 1 if "Protocolo:" in headers else None
    col_valor_total = headers.index(
        "Valor da saida:") + 1 if "Valor da saida:" in headers else None
    col_km = headers.index("Adicional de KM:") + \
        1 if "Adicional de KM:" in headers else None
    col_total = headers.index("Total:") + 1 if "Total:" in headers else None

    # Find the correct row to insert data
    insert_row = find_insert_row(sheet)

    # Insert a new row instead of overwriting
    sheet.insert_rows(insert_row)

    # Define font size 8
    font_style = Font(size=8)

    # Insert new data in the correct columns and apply font size
    if col_cliente:
        cell = sheet.cell(row=insert_row, column=col_cliente,
                          value=data['Cliente'])
        cell.font = font_style
    if col_protocolo:
        cell = sheet.cell(row=insert_row, column=col_protocolo,
                          value=data['Protocolo'])
        cell.font = font_style
    if col_valor_total:
        valor = float(data['Valor Total'].replace(
            ',', '.')) if data['Valor Total'] else 0.0
        cell = sheet.cell(
            row=insert_row, column=col_valor_total, value=f"R$ {valor:.2f}")
        cell.font = font_style
    if col_km:
        cell = sheet.cell(row=insert_row, column=col_km, value=data['KM'])
        cell.font = font_style
    if col_total:
        cell = sheet.cell(row=insert_row, column=col_total,
                          value=f"R$ {valor:.2f}")
        cell.font = font_style

    wb.save(filename)
    print(f"Data successfully added to {filename}")


if __name__ == "__main__":
    file_path = "input.txt"
    excel_path = "Rota 24 Horas.xlsx"

    try:
        with open(file_path, "r", encoding="utf-8") as file:
            text = file.read()

        extracted_data = extract_data(text)
        write_to_excel(extracted_data, excel_path)
    except FileNotFoundError:
        print("File not found. Please check the file path and try again.")
